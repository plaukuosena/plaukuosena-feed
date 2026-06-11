#!/usr/bin/env python3
"""
Plaukuosena.lt -> Kaina24 XML feed generator.

Generates:
- kaina24.xml
- products_snapshot.csv
- products_snapshot.xlsx

Included fixes:
- UTF-8 Lithuanian letters
- Kaina24 XML structure
- EAN extraction from product page content
- Manufacturer / brand extraction from product title
"""

from __future__ import annotations

import csv
import hashlib
import html
import json
import re
import sys
import time
import unicodedata
import xml.etree.ElementTree as ET
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable

import requests
requests.packages.urllib3.disable_warnings()

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / "config.json"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "lt-LT,lt;q=0.9,en-US;q=0.8,en;q=0.7",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
}


@dataclass
class Product:
    sku: str
    title: str
    brand: str
    price: str
    old_price: str
    currency: str
    url: str
    image: str
    category: str
    availability: str
    size: str
    description: str
    ean_code: str
    source_status: str


def load_config() -> dict[str, Any]:
    with CONFIG_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def fetch_text(url: str, timeout: int = 25) -> str:
    session = requests.Session()

    response = session.get(
        url,
        headers=HEADERS,
        timeout=timeout,
        verify=False
    )

    if response.status_code == 403:
        alt_headers = HEADERS.copy()
        alt_headers["Referer"] = "https://www.google.com/"

        response = session.get(
            url,
            headers=alt_headers,
            timeout=timeout,
            verify=False
        )

    response.raise_for_status()
    response.encoding = "utf-8"

    return response.text


def normalize_text(value: str | None) -> str:
    if not value:
        return ""

    value = html.unescape(value)
    value = re.sub(r"\s+", " ", value).strip()

    return value


def strip_accents(value: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(ch)
    )


def codeify(value: str, max_len: int = 12) -> str:
    value = strip_accents(value).upper()
    value = re.sub(r"[^A-Z0-9]+", "-", value).strip("-")
    value = re.sub(r"-+", "-", value)

    return value[:max_len].strip("-") or "PROD"


def sitemap_urls(xml_text: str) -> list[str]:
    soup = BeautifulSoup(xml_text, "xml")
    urls: list[str] = []

    for loc in soup.find_all("loc"):
        if loc.text:
            urls.append(loc.text.strip())

    return urls


def likely_product_url(url: str, cfg: dict[str, Any]) -> bool:
    filters = cfg.get("filters", {})
    low = url.lower()

    for bad in filters.get("exclude_url_contains", []):
        if bad.lower() in low:
            return False

    includes = filters.get("include_url_contains", [])

    if includes and not any(x.lower() in low for x in includes):
        return False

    hints = filters.get("product_hint_words", [])

    return any(h.lower() in low for h in hints)


def jsonld_objects(soup: BeautifulSoup) -> Iterable[Any]:
    for script in soup.find_all("script", type="application/ld+json"):
        raw = script.string or script.get_text(" ", strip=True)

        if not raw:
            continue

        try:
            data = json.loads(raw)
        except Exception:
            continue

        if isinstance(data, list):
            for item in data:
                yield item
        else:
            yield data


def iter_graph_nodes(obj: Any) -> Iterable[dict[str, Any]]:
    if isinstance(obj, dict):
        if "@graph" in obj and isinstance(obj["@graph"], list):
            for node in obj["@graph"]:
                if isinstance(node, dict):
                    yield from iter_graph_nodes(node)
        else:
            yield obj

    elif isinstance(obj, list):
        for item in obj:
            yield from iter_graph_nodes(item)


def find_product_jsonld(soup: BeautifulSoup) -> dict[str, Any] | None:
    for obj in jsonld_objects(soup):
        for node in iter_graph_nodes(obj):
            typ = node.get("@type")

            if isinstance(typ, list):
                is_product = "Product" in typ
            else:
                is_product = typ == "Product"

            if is_product:
                return node

    return None


def get_meta(soup: BeautifulSoup, *names: str) -> str:
    for name in names:
        tag = (
            soup.find("meta", attrs={"property": name})
            or soup.find("meta", attrs={"name": name})
        )

        if tag and tag.get("content"):
            return normalize_text(tag.get("content"))

    return ""


def first_image_from_json(value: Any) -> str:
    if isinstance(value, str):
        return value

    if isinstance(value, list) and value:
        return first_image_from_json(value[0])

    if isinstance(value, dict):
        return value.get("url", "") or value.get("contentUrl", "")

    return ""


def offer_data(product_ld: dict[str, Any]) -> dict[str, Any]:
    offers = product_ld.get("offers") or {}

    if isinstance(offers, list) and offers:
        offers = offers[0]

    if not isinstance(offers, dict):
        return {}

    return offers


def infer_brand(title: str, cfg: dict[str, Any], ld_brand: Any = None) -> str:
    if isinstance(ld_brand, dict):
        name = ld_brand.get("name")

        if name:
            return normalize_text(str(name))

    if isinstance(ld_brand, str) and ld_brand:
        return normalize_text(ld_brand)

    known_brands = [
        "WELLA PROFESSIONALS",
        "Wella Professionals",
        "American Crew",
        "AMERICAN CREW",
        "MILK SHAKE",
        "Milk Shake",
        "Kadus Professional",
        "Kadus Professionals",
        "KADUS PROFESSIONALS",
        "OLAPLEX",
        "Olaplex",
        "DAVINES",
        "Davines",
        "QIQI",
        "Qiqi",
        "L'Oréal Professionnel",
        "Loreal Professionnel",
        "L'Oreal Professionnel",
        "L'OREAL PROFESSIONNEL",
    ]

    low_title = title.lower()

    for brand in known_brands:
        if low_title.startswith(brand.lower()):
            return brand

    for brand in known_brands:
        if brand.lower() in low_title:
            return brand

    first_word = title.strip().split(" ")[0]

    return first_word if first_word else ""


def infer_size(text: str) -> str:
    match = re.search(
        r"(\d+(?:[,.]\d+)?)\s?(ml|g|l|kg|vnt|pcs)\b",
        text,
        flags=re.I
    )

    if not match:
        return ""

    number = match.group(1).replace(",", ".")
    unit = match.group(2).lower()

    return f"{number}{unit}"


def infer_type_code(title: str, cfg: dict[str, Any]) -> str:
    low = title.lower()

    for key, code in cfg.get("type_codes", {}).items():
        if key.lower() in low:
            return code

    return "PRD"


def model_code(title: str, brand: str, type_code: str) -> str:
    clean = title

    if brand:
        clean = re.sub(re.escape(brand), " ", clean, flags=re.I)

    common = [
        "shampoo", "conditioner", "mask", "oil", "spray", "serum", "cream", "gel",
        "šampūnas", "sampunas", "kondicionierius", "kaukė", "kauke", "aliejus",
        "purškiklis", "purskiklis", "serumas", "kremas", "gelis",
        "plaukams", "plauku", "plaukų", "prieziuros", "priežiūros"
    ]

    for word in common:
        clean = re.sub(rf"\b{re.escape(word)}\b", " ", clean, flags=re.I)

    clean = re.sub(
        r"\d+(?:[,.]\d+)?\s?(ml|g|l|kg|vnt|pcs)\b",
        " ",
        clean,
        flags=re.I
    )

    tokens = [
        codeify(t, 8)
        for t in re.split(r"\s+", clean)
        if codeify(t, 8)
    ]

    tokens = [
        t for t in tokens
        if t not in {type_code, "PROD", "IR", "SU", "BE", "THE"}
    ]

    if not tokens:
        return "GEN"

    return "-".join(tokens[:2])


def generate_sku(title: str, brand: str, size: str, cfg: dict[str, Any]) -> str:
    brand_code = cfg.get("brand_codes", {}).get(brand) or codeify(brand or "PLK", 4)
    type_code = infer_type_code(title, cfg)
    model = model_code(title, brand, type_code)
    size_code = codeify(size.replace(".", ""), 8) if size else "NA"
    base = f"{brand_code}-{model}-{type_code}-{size_code}"

    return re.sub(r"-+", "-", base).strip("-")


def availability_from_offer(offer: dict[str, Any], cfg: dict[str, Any]) -> str:
    av = str(offer.get("availability", "")).lower()

    if "instock" in av or "in_stock" in av or "in stock" in av:
        return "in_stock"

    if "outofstock" in av or "out_of_stock" in av or "out of stock" in av:
        return "out_of_stock"

    if "preorder" in av:
        return "preorder"

    return cfg.get("store", {}).get("default_availability", "in_stock")


def clean_price(value: Any) -> str:
    if value is None:
        return ""

    text = str(value)
    text = text.replace("€", "").replace("EUR", "")
    text = re.sub(r"[^0-9,.]", "", text).replace(",", ".")

    match = re.search(r"\d+(?:\.\d+)?", text)

    return match.group(0) if match else ""


def kaina24_category(title: str) -> str:
    low = title.lower()

    if "šampūn" in low or "sampun" in low or "shampoo" in low:
        return "Plaukų šampūnai"

    if "kondicionier" in low or "conditioner" in low:
        return "Plaukų kondicionieriai"

    if "kauk" in low or "mask" in low:
        return "Plaukų kaukės"

    if "aliej" in low or "oil" in low:
        return "Plaukų aliejai"

    if "puršk" in low or "pursk" in low or "spray" in low:
        return "Plaukų purškikliai"

    if "serum" in low:
        return "Plaukų serumai"

    if "gel" in low:
        return "Plaukų geliai"

    return "Plaukų priežiūros priemonės"


def extract_ean_from_page(page_html: str) -> str:
    if not page_html:
        return ""

    decoded = html.unescape(page_html)

    plain_text = BeautifulSoup(decoded, "lxml").get_text(" ", strip=True)

    # Search for the closest 8-14 digit number after the EAN label.
    label_match = re.search(
        r"\bEAN\b(.{0,2500}?)(\d{8,14})",
        plain_text,
        flags=re.I | re.S
    )

    if label_match:
        return label_match.group(2)

    # Fallback for encoded HTML/source.
    html_match = re.search(
        r"EAN.{0,5000}?(\d{8,14})",
        decoded,
        flags=re.I | re.S
    )

    if html_match:
        return html_match.group(1)

    return ""


def parse_product(url: str, cfg: dict[str, Any]) -> Product | None:
    try:
        page = fetch_text(url)
    except Exception as exc:
        print(f"WARN fetch failed: {url} ({exc})", file=sys.stderr)
        return None

    soup = BeautifulSoup(page, "lxml")
    ld = find_product_jsonld(soup) or {}
    offer = offer_data(ld)

    title = normalize_text(ld.get("name") if isinstance(ld, dict) else "")

    if not title:
        title = get_meta(soup, "og:title", "twitter:title")

    if not title and soup.title:
        title = normalize_text(soup.title.get_text(" ", strip=True))

    title = re.sub(
        r"\s*[|–-]\s*Plaukuosena.*$",
        "",
        title,
        flags=re.I
    ).strip()

    if not title:
        return None

    brand = infer_brand(
        title,
        cfg,
        ld.get("brand") if isinstance(ld, dict) else None
    )

    description = normalize_text(
        ld.get("description") if isinstance(ld, dict) else ""
    )

    if not description:
        description = get_meta(soup, "og:description", "description")

    image = first_image_from_json(
        ld.get("image") if isinstance(ld, dict) else None
    )

    if not image:
        image = get_meta(soup, "og:image", "twitter:image")

    price = clean_price(
        offer.get("price") or get_meta(soup, "product:price:amount")
    )

    currency = normalize_text(
        offer.get("priceCurrency")
        or get_meta(soup, "product:price:currency")
        or cfg["store"].get("currency", "EUR")
    )

    availability = availability_from_offer(offer, cfg)
    size = infer_size(title + " " + description)
    category = kaina24_category(title)
    ean_code = extract_ean_from_page(page)

    sku = normalize_text(
        ld.get("sku") if isinstance(ld, dict) else ""
    )

    if not sku:
        sku = generate_sku(title, brand, size, cfg)

    return Product(
        sku=sku,
        title=title,
        brand=brand,
        price=price,
        old_price="",
        currency=currency,
        url=url,
        image=image,
        category=category,
        availability=availability,
        size=size,
        description=description[:500],
        ean_code=ean_code,
        source_status="OK" if price else "CHECK_PRICE"
    )


def dedupe_skus(products: list[Product]) -> list[Product]:
    seen: dict[str, int] = {}

    for product in products:
        base = product.sku

        if base not in seen:
            seen[base] = 1
            continue

        seen[base] += 1
        suffix = hashlib.md5(product.url.encode("utf-8")).hexdigest()[:4].upper()
        product.sku = f"{base}-{suffix}"

    return products


def write_csv(products: list[Product], path: Path) -> None:
    fields = list(asdict(products[0]).keys()) if products else [
        "sku",
        "title",
        "brand",
        "price",
        "old_price",
        "currency",
        "url",
        "image",
        "category",
        "availability",
        "size",
        "description",
        "ean_code",
        "source_status"
    ]

    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()

        for product in products:
            writer.writerow(asdict(product))


def write_xlsx(products: list[Product], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "sku",
        "title",
        "brand",
        "price",
        "currency",
        "url",
        "image",
        "category",
        "availability",
        "size",
        "ean_code",
        "source_status"
    ]

    ws.append(headers)

    for product in products:
        ws.append([getattr(product, h) for h in headers])

    header_fill = PatternFill("solid", fgColor="0F766E")

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = {
        "A": 22,
        "B": 48,
        "C": 22,
        "D": 10,
        "E": 10,
        "F": 55,
        "G": 45,
        "H": 22,
        "I": 14,
        "J": 16,
        "K": 16,
        "L": 16,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    wb.save(path)


def cdata(text: str) -> str:
    text = text or ""
    text = text.replace("]]>", "]]]]><![CDATA[>")

    return f"<![CDATA[ {text} ]]>"


def write_kaina24_xml(products: list[Product], path: Path, cfg: dict[str, Any]) -> None:
    lines = ['<?xml version="1.0" encoding="UTF-8"?>', '<products>']

    for product in products:
        lines.append(f'  <product id="{product.sku}">')
        lines.append(f'    <title>{cdata(product.title)}</title>')
        lines.append(f'    <description>{cdata(product.description)}</description>')
        lines.append(f'    <price>{product.price}</price>')

        if product.ean_code:
            lines.append(f'    <ean_code>{cdata(product.ean_code)}</ean_code>')

        lines.append('    <condition>new</condition>')
        lines.append('    <stock>5</stock>')
        lines.append(f'    <manufacturer>{cdata(product.brand)}</manufacturer>')
        lines.append(f'    <model>{cdata(product.sku)}</model>')
        lines.append(f'    <image_url>{cdata(product.image)}</image_url>')
        lines.append(f'    <product_url>{cdata(product.url)}</product_url>')
        lines.append(f'    <category_name>{cdata(product.category)}</category_name>')
        lines.append('    <delivery>')
        lines.append('      <home_delivery>')
        lines.append('        <working_days><![CDATA[ 2 ]]></working_days>')
        lines.append('        <price><![CDATA[ 3.99 ]]></price>')
        lines.append('      </home_delivery>')
        lines.append('    </delivery>')
        lines.append('  </product>')

    lines.append('</products>')

    path.write_text("\n".join(lines), encoding="utf-8")

def product_urls_from_kaina24_xml() -> list[str]:
    path = ROOT / "kaina24.xml"
    if not path.exists():
        return []

    try:
        tree = ET.parse(path)
        root = tree.getroot()
        urls = []

        for el in root.iter():
            if el.tag.endswith("product_url") and el.text:
                url = el.text.strip()
                if url:
                    urls.append(url)

        return list(dict.fromkeys(urls))
    except Exception:
        return []

def main() -> int:
    cfg = load_config()

sitemap = None
try:
    sitemap = fetch_text(cfg["store"]["sitemap_url"])
except Exception:
    sitemap = None

if sitemap:
    all_urls = sitemap_urls(sitemap)
    product_urls = [url for url in all_urls if likely_product_url(url, cfg)]
else:
    product_urls = product_urls_from_kaina24_xml()
    all_urls = product_urls

    print(
        f"Found {len(all_urls)} sitemap URLs; "
        f"{len(product_urls)} likely product URLs"
    )

    products: list[Product] = []

    for i, url in enumerate(product_urls, 1):
        print(f"[{i}/{len(product_urls)}] {url}")

        product = parse_product(url, cfg)

        if product:
            products.append(product)

        time.sleep(0.25)

    products = dedupe_skus(products)
    products.sort(key=lambda x: (x.brand, x.title))

    out_cfg = cfg.get("output", {})

    write_kaina24_xml(
        products,
        ROOT / out_cfg.get("xml_file", "kaina24.xml"),
        cfg
    )

    write_csv(
        products,
        ROOT / out_cfg.get("csv_file", "products_snapshot.csv")
    )

    write_xlsx(
        products,
        ROOT / out_cfg.get("xlsx_file", "products_snapshot.xlsx")
    )

    print(f"Generated {len(products)} products")
    print(f"XML: {ROOT / out_cfg.get('xml_file', 'kaina24.xml')}")

    return 0

if __name__ == "__main__":
    raise SystemExit(main())
